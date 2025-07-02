from flask import Flask, render_template, request, redirect, url_for, session, flash
import xlsxwriter
from datetime import datetime
import xml.etree.ElementTree as ET
from flask import make_response
import os
import io
import xml.etree.ElementTree as ET
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import pandas as pd
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import json


def get_db_connection():
    conn = sqlite3.connect('training_management.db')
    conn.row_factory = sqlite3.Row
    return conn


app = Flask(__name__)
app.secret_key = '3315'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xml', 'xlsx'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.jinja_env.add_extension('jinja2.ext.do')

@app.template_filter('fromjson')
def fromjson_filter(value):
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value

@app.template_filter('tojson')
def tojson_filter(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('role_selection'))

            if role and session.get('role') != role:
                return redirect(url_for('logout'))

            return f(*args, **kwargs)

        return decorated_function

    return decorator


def format_datetime(value, format='%Y-%m-%d %H:%M'):
    if value is None:
        return ''
    if isinstance(value, str):
        try:
            value = datetime.strptime(value, '%Y-%m-%d %H:%M')
        except ValueError:
            return value
    return value.strftime(format)


app.jinja_env.filters['datetimeformat'] = format_datetime


@app.context_processor
def inject_datetime():
    return {'datetime': datetime}


@app.route('/')
def role_selection():
    return render_template('auth/login.html')


@app.route('/login/<role>', methods=['GET', 'POST'])
def login(role):
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute(
            'SELECT * FROM users WHERE username = ? AND role = ?',
            (username, role)
        ).fetchone()
        conn.close()

        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            if user['department_id']:
                conn = get_db_connection()
                dept = conn.execute(
                    'SELECT name FROM departments WHERE id = ?',
                    (user['department_id'],)
                ).fetchone()
                conn.close()

                if dept:
                    session['department'] = dept['name']
                    session['department_id'] = user['department_id']
            return redirect(url_for('dashboard'))

        return render_template('auth/credentials.html', role=role, error="Invalid credentials")

    return render_template('auth/credentials.html', role=role)


@app.route('/dashboard')
def dashboard():
    role = session.get('role')
    if not role:
        return redirect(url_for('role_selection'))

    if role == 'ld_admin':
        return render_template('L&D/dashboard.html')
    elif role.endswith('_admin'):
        return render_template('Department/dashboard_dept.html', department=session.get('department'))
    else:
        return redirect(url_for('logout'))


@app.route('/upload', methods=['GET', 'POST'])
@login_required(role='ld_admin')
def upload_file():
    if request.method == 'POST':
        if 'training_file' not in request.files:
            flash('No file part')
            return redirect(request.url)

        file = request.files['training_file']
        department = request.form.get('department')

        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)

        if not department:
            flash('No department selected')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(f"{department}_training_format.{file.filename.rsplit('.', 1)[1].lower()}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            try:
                file.save(filepath)

                conn = get_db_connection()
                dept = conn.execute(
                    'SELECT id FROM departments WHERE name = ?',
                    (department,)
                ).fetchone()

                if not dept:
                    os.remove(filepath)
                    flash('Invalid department', 'error')
                    return redirect(request.url)

                field_names = []
                if filename.endswith('.xml'):
                    tree = ET.parse(filepath)
                    root = tree.getroot()
                    if len(root) > 0:
                        field_names = [elem.tag for elem in root[0]]
                    else:
                        os.remove(filepath)
                        flash('XML file has no content', 'error')
                        return redirect(request.url)
                elif filename.endswith('.xlsx'):
                    wb = load_workbook(filepath)
                    sheet = wb.active
                    field_names = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

                conn.execute(
                    'INSERT INTO training_formats (department_id, file_name, file_path, file_type, field_names, uploaded_by) '
                    'VALUES (?, ?, ?, ?, ?, ?)',
                    (dept['id'], filename, filepath, filename.split('.')[-1],
                     json.dumps(field_names), session['user_id'])
                )
                conn.commit()
                conn.close()

                flash('File successfully uploaded', 'success')
                return redirect(url_for('dashboard'))
            except Exception as e:
                if os.path.exists(filepath):
                    os.remove(filepath)
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)

    flash('Invalid file type. Only XML or XLSX files are allowed', 'error')
    return redirect(request.url)


@app.route('/upload_page')
def upload_page():
    return render_template('upload.html')


@app.route('/dept_training_content', methods=['GET', 'POST'])
@login_required(role='dept_admin')
def dept_training_content():
    department_id = session.get('department_id')

    if request.method == 'POST':
        training_data = request.form.to_dict()
        assigned_employees = request.form.getlist('assigned_employees')

        if not training_data.get('training_name'):
            flash('Training name is required', 'error')
            return redirect(url_for('dept_training_content'))

        conn = get_db_connection()

        try:
            cursor = conn.execute(
                'INSERT INTO training_programs (department_id, training_name, training_data, created_by) '
                'VALUES (?, ?, ?, ?)',
                (department_id, training_data['training_name'],
                 json.dumps(training_data), session['user_id'])
            )
            training_id = cursor.lastrowid

            for emp_id in assigned_employees:
                conn.execute(
                    'INSERT INTO training_assignments (training_id, employee_id, assigned_by) '
                    'VALUES (?, ?, ?)',
                    (training_id, emp_id, session['user_id'])
                )

            conn.commit()
            conn.close()
            flash('Training content saved successfully!', 'success')
        except Exception as e:
            conn.close()
            flash(f'Error saving training: {str(e)}', 'error')

        return redirect(url_for('dept_training_content'))

    conn = get_db_connection()

    format_data = conn.execute(
        'SELECT * FROM training_formats WHERE department_id = ? ORDER BY created_at DESC LIMIT 1',
        (department_id,)
    ).fetchone()

    employees = conn.execute(
        'SELECT * FROM employees WHERE department_id = ?',
        (department_id,)
    ).fetchall()

    trainings = conn.execute(
        'SELECT * FROM training_programs WHERE department_id = ? ORDER BY created_at DESC',
        (department_id,)
    ).fetchall()

    trainings_list = [dict(training) for training in trainings]
    employees_list = [dict(employee) for employee in employees]
    format_data_dict = dict(format_data) if format_data else None

    conn.close()

    field_names = json.loads(format_data_dict['field_names']) if format_data_dict else []

    return render_template('Department/training_content.html',
                           department=session.get('department'),
                           field_names=field_names,
                           existing_values={},
                           employees=employees_list,
                           training_records=trainings_list,
                           datetime=datetime)


@app.route('/edit_training/<int:record_id>', methods=['GET', 'POST'])
@login_required(role='dept_admin')
def edit_training(record_id):
    if 'department_id' not in session:
        return redirect(url_for('logout'))

    conn = get_db_connection()

    try:
        training = conn.execute(
            'SELECT * FROM training_programs WHERE id = ? AND department_id = ?',
            (record_id, session['department_id'])
        ).fetchone()

        if not training:
            flash('Training record not found', 'error')
            return redirect(url_for('dept_training_content'))

        training_data = json.loads(training['training_data'])

        format_data = conn.execute(
            'SELECT field_names FROM training_formats WHERE department_id = ? ORDER BY created_at DESC LIMIT 1',
            (session['department_id'],)
        ).fetchone()

        field_names = json.loads(format_data['field_names']) if format_data else []

        employees = conn.execute(
            'SELECT * FROM employees WHERE department_id = ?',
            (session['department_id'],)
        ).fetchall()
        assigned_employees = conn.execute(
            'SELECT employee_id FROM training_assignments WHERE training_id = ?',
            (record_id,)
        ).fetchall()
        assigned_ids = [str(emp['employee_id']) for emp in assigned_employees]

        if request.method == 'POST':
            # Update training data from form
            updated_data = request.form.to_dict()
            new_assigned = request.form.getlist('assigned_employees')

            for field in training_data:
                if field not in updated_data and field != 'assigned_employees':
                    updated_data[field] = training_data[field]

            conn.execute(
                'UPDATE training_programs SET training_name = ?, training_data = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (updated_data.get('training_name', training['training_name']),
                 json.dumps(updated_data),
                 record_id
                 ))

            conn.execute('DELETE FROM training_assignments WHERE training_id = ?', (record_id,))
            for emp_id in new_assigned:
                conn.execute(
                    'INSERT INTO training_assignments (training_id, employee_id, assigned_by) VALUES (?, ?, ?)',
                    (record_id, emp_id, session['user_id'])
                )

            conn.commit()
            flash('Training record updated successfully!', 'success')
            return redirect(url_for('dept_training_content'))

        return render_template('Department/edit_training.html',
                               department=session.get('department'),
                               field_names=field_names,
                               record=training_data,
                               employees=employees,
                               assigned_employees=assigned_ids)
    except Exception as e:
        conn.rollback()
        flash(f'Error updating training: {str(e)}', 'error')
        return redirect(url_for('dept_training_content'))
    finally:
        conn.close()

@app.route('/delete_training/<record_id>')
def delete_training(record_id):
    if 'department_id' not in session:
        return redirect(url_for('logout'))

    conn = get_db_connection()

    try:
        training = conn.execute(
            'SELECT id FROM training_programs WHERE id = ? AND department_id = ?',
            (record_id, session['department_id'])
        ).fetchone()

        if not training:
            flash('Training record not found or not authorized', 'error')
            return redirect(url_for('dept_training_content'))

        conn.execute('DELETE FROM training_programs WHERE id = ?', (record_id,))
        conn.commit()
        flash('Training record deleted successfully', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting training record: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('dept_training_content'))


@app.route('/export_training/<record_id>')
def export_training(record_id):
    conn = get_db_connection()

    training = conn.execute(
        'SELECT * FROM training_programs WHERE id = ?',
        (record_id,)
    ).fetchone()

    if not training:
        conn.close()
        flash('Training record not found', 'error')
        return redirect(url_for('dept_training_content'))

    employees = conn.execute('''
        SELECT e.name, e.employee_id, e.designation
        FROM training_assignments ta
        JOIN employees e ON ta.employee_id = e.id
        WHERE ta.training_id = ?
    ''', (record_id,)).fetchall()

    format_data = conn.execute(
        'SELECT field_names FROM training_formats WHERE department_id = ? ORDER BY created_at DESC LIMIT 1',
        (training['department_id'],)
    ).fetchone()
    conn.close()

    field_names = json.loads(format_data['field_names']) if format_data else []
    training_data = json.loads(training['training_data'])

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Training Program')

    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    bold_format = workbook.add_format({'bold': True, 'border': 1})
    regular_format = workbook.add_format({'border': 1})
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})

    worksheet.write(0, 0, 'Training Name:', bold_format)
    worksheet.write(0, 1, training['training_name'], regular_format)
    worksheet.write(1, 0, 'Created At:', bold_format)
    worksheet.write(1, 1, training['created_at'], date_format)

    row = 3
    for field in field_names:
        if field != 'training_name':
            worksheet.write(row, 0, f"{field.replace('_', ' ').title()}:", bold_format)
            worksheet.write(row, 1, training_data.get(field, 'N/A'), regular_format)
            row += 1

    row += 1
    worksheet.write(row, 0, 'Assigned Employees', header_format)
    worksheet.write(row, 1, 'Employee ID', header_format)
    worksheet.write(row, 2, 'Designation', header_format)
    row += 1

    for emp in employees:
        worksheet.write(row, 0, emp['name'], regular_format)
        worksheet.write(row, 1, emp['employee_id'], regular_format)
        worksheet.write(row, 2, emp['designation'] or 'N/A', regular_format)
        row += 1

    workbook.close()
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=training_program_{record_id}.xlsx'
    return response


@app.route('/export_period')
def export_by_period():
    department = request.args.get('department')
    month = request.args.get('month')
    year = request.args.get('year')

    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        flash('Department not found', 'error')
        return redirect(url_for('dept_training_content'))

    department_id = dept['id']

    query_params = [department_id]
    date_filter = ""

    if month:
        date_filter += " AND strftime('%m', tp.created_at) = ?"
        query_params.append(f"{int(month):02d}")
    if year:
        date_filter += " AND strftime('%Y', tp.created_at) = ?"
        query_params.append(year)

    trainings = conn.execute(f'''
        SELECT tp.*, GROUP_CONCAT(e.name, ', ') as assigned_employees
        FROM training_programs tp
        LEFT JOIN training_assignments ta ON tp.id = ta.training_id
        LEFT JOIN employees e ON ta.employee_id = e.id
        WHERE tp.department_id = ? {date_filter}
        GROUP BY tp.id
        ORDER BY tp.created_at DESC
    ''', query_params).fetchall()

    format_data = conn.execute(
        'SELECT field_names FROM training_formats WHERE department_id = ? ORDER BY created_at DESC LIMIT 1',
        (department_id,)
    ).fetchone()
    conn.close()

    field_names = json.loads(format_data['field_names']) if format_data else []

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Training Programs')

    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})
    regular_format = workbook.add_format({'border': 1})

    headers = ['Training Name', 'Created At'] + [f.replace('_', ' ').title() for f in field_names if
                                                 f != 'training_name'] + ['Assigned Employees']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    for row, training in enumerate(trainings, start=1):
        training_data = json.loads(training['training_data'])

        worksheet.write(row, 0, training['training_name'], regular_format)

        worksheet.write(row, 1, training['created_at'], date_format)

        col = 2
        for field in field_names:
            if field != 'training_name':
                worksheet.write(row, col, training_data.get(field, 'N/A'), regular_format)
                col += 1

        worksheet.write(row, col, training['assigned_employees'], regular_format)

    workbook.close()
    output.seek(0)

    filename = f"training_programs_{department}"
    if year:
        filename += f"_{year}"
    if month:
        filename += f"_{month}"
    filename += ".xlsx"

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/export_all')
def export_all():
    department = request.args.get('department')
    month = request.args.get('month')
    year = request.args.get('year')

    # Get department ID
    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        flash('Department not found', 'error')
        return redirect(url_for('dept_training_content'))

    department_id = dept['id']

    query_params = [department_id]
    date_filter = ""

    if month:
        date_filter += " AND strftime('%m', tp.created_at) = ?"
        query_params.append(f"{int(month):02d}")
    if year:
        date_filter += " AND strftime('%Y', tp.created_at) = ?"
        query_params.append(year)

    trainings = conn.execute(f'''
        SELECT tp.*, GROUP_CONCAT(e.name, ', ') as assigned_employees
        FROM training_programs tp
        LEFT JOIN training_assignments ta ON tp.id = ta.training_id
        LEFT JOIN employees e ON ta.employee_id = e.id
        WHERE tp.department_id = ? {date_filter}
        GROUP BY tp.id
        ORDER BY tp.created_at DESC
    ''', query_params).fetchall()

    format_data = conn.execute(
        'SELECT field_names FROM training_formats WHERE department_id = ? ORDER BY created_at DESC LIMIT 1',
        (department_id,)
    ).fetchone()
    conn.close()

    field_names = json.loads(format_data['field_names']) if format_data else []

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Training Programs')

    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})
    regular_format = workbook.add_format({'border': 1})

    headers = ['Training Name', 'Created At'] + [f.replace('_', ' ').title() for f in field_names if
                                                 f != 'training_name'] + ['Assigned Employees']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Write data
    for row, training in enumerate(trainings, start=1):
        training_data = json.loads(training['training_data'])

        worksheet.write(row, 0, training['training_name'], regular_format)

        worksheet.write(row, 1, training['created_at'], date_format)

        col = 2
        for field in field_names:
            if field != 'training_name':
                worksheet.write(row, col, training_data.get(field, 'N/A'), regular_format)
                col += 1

        worksheet.write(row, col, training['assigned_employees'], regular_format)

    workbook.close()
    output.seek(0)

    filename = f"all_training_programs_{department}.xlsx"

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/lnddepartments')
def lnddepartments():
    return render_template('L&D/departments.html')



@app.route('/dept_att')
def dept_att():
    department = session.get('department')
    if not department:
        return redirect(url_for('logout'))

    conn = get_db_connection()

    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        return redirect(url_for('logout'))

    department_id = dept['id']

    trainings = conn.execute(
        'SELECT * FROM training_programs WHERE department_id = ?',
        (department_id,)
    ).fetchall()

    attendance_records = conn.execute('''
        SELECT ar.id as record_id, ar.session_date, tp.training_name, 
               GROUP_CONCAT(e.name, ', ') as participants
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        LEFT JOIN attendance_details ad ON ar.id = ad.attendance_id
        LEFT JOIN employees e ON ad.employee_id = e.id
        WHERE tp.department_id = ?
        GROUP BY ar.id
        ORDER BY ar.session_date DESC
    ''', (department_id,)).fetchall()

    employees = conn.execute(
        'SELECT * FROM employees WHERE department_id = ?',
        (department_id,)
    ).fetchall()

    conn.close()

    return render_template('Department/attendance.html',
                         department=department,
                         employees=employees,
                         trainings=trainings,
                         attendance_records=attendance_records,
                         current_date=datetime.now().strftime('%Y-%m-%d'))


@app.route('/add_employee', methods=['GET', 'POST'])
@login_required(role='dept_admin')
def add_employee():
    department = session.get('department')
    if not department:
        return redirect(url_for('logout'))

    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()
    conn.close()

    if not dept:
        flash('Department not found', 'error')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        employee_id = request.form.get('employee_id')
        designation = request.form.get('designation')

        try:
            conn = get_db_connection()
            conn.execute(
                'INSERT INTO employees (employee_id, name, email, designation, department_id) '
                'VALUES (?, ?, ?, ?, ?)',
                (employee_id, name, email, designation, dept['id'])
            )
            conn.commit()
            conn.close()
            flash('Employee added successfully!', 'success')
            return redirect(url_for('view_employees'))
        except sqlite3.IntegrityError as e:
            conn.close()
            flash('Error adding employee: ID or email already exists', 'error')

    return render_template('Department/add_employee.html', department=department)


@app.route('/view_employees')
@login_required(role='dept_admin')
def view_employees():
    department = session.get('department')
    if not department:
        return redirect(url_for('logout'))

    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        flash('Department not found', 'error')
        return redirect(url_for('dashboard'))

    employees = conn.execute(
        'SELECT * FROM employees WHERE department_id = ?',
        (dept['id'],)
    ).fetchall()
    conn.close()

    return render_template('Department/view_employees.html',
                           employees=employees,
                           department=department)


@app.route('/edit_employee/<int:employee_id>', methods=['GET', 'POST'])
@login_required(role='dept_admin')
def edit_employee(employee_id):
    conn = get_db_connection()
    employee = conn.execute(
        'SELECT * FROM employees WHERE id = ? AND department_id = ?',
        (employee_id, session.get('department_id'))
    ).fetchone()

    if not employee:
        conn.close()
        flash('Employee not found', 'error')
        return redirect(url_for('view_employees'))

    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        designation = request.form.get('designation')

        try:
            conn.execute(
                'UPDATE employees SET name = ?, email = ?, designation = ?, updated_at = CURRENT_TIMESTAMP '
                'WHERE id = ?',
                (name, email, designation, employee_id)
            )
            conn.commit()
            conn.close()
            flash('Employee updated successfully!', 'success')
            return redirect(url_for('view_employees'))
        except sqlite3.IntegrityError:
            conn.close()
            flash('Error updating employee: email already exists', 'error')

    conn.close()
    return render_template('Department/edit_employee.html', employee=employee)

@app.route('/delete_employee/<int:employee_id>')
@login_required(role='dept_admin')
def delete_employee(employee_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM employees WHERE id = ?', (employee_id,))
    conn.commit()
    conn.close()
    flash('Employee deleted successfully', 'success')
    return redirect(url_for('view_employees'))


@app.route('/mark_attendance', methods=['POST'])
@login_required(role='dept_admin')
def mark_attendance():
    department_id = session.get('department_id')
    training_id = request.form.get('training_name')
    attendance_date = request.form.get('attendance_date')
    present_employee_ids = request.form.getlist('present')

    conn = get_db_connection()

    try:
        training = conn.execute(
            'SELECT id FROM training_programs WHERE id = ? AND department_id = ?',
            (training_id, department_id)
        ).fetchone()

        if not training:
            flash('Training record not found', 'error')
            return redirect(url_for('attendance'))

        cursor = conn.execute(
            'INSERT INTO attendance_records (training_id, session_date, conducted_by) '
            'VALUES (?, ?, ?)',
            (training_id, attendance_date, session['user_id'])
        )
        attendance_id = cursor.lastrowid

        for emp_id in present_employee_ids:
            conn.execute(
                'INSERT INTO attendance_details (attendance_id, employee_id) '
                'VALUES (?, ?)',
                (attendance_id, emp_id)
            )

        conn.commit()
        conn.close()
        flash('Attendance saved successfully!', 'success')
    except Exception as e:
        conn.close()
        flash(f'Error saving attendance: {str(e)}', 'error')

    return redirect(url_for('attendance'))

@app.route('/attendance', methods=['GET'])
def attendance():
    if 'department' not in session:
        return redirect(url_for('logout'))

    department = session['department']
    department_id = session.get('department_id')
    today = datetime.now().strftime('%Y-%m-%d')

    conn = get_db_connection()

    employees = conn.execute(
        'SELECT * FROM employees WHERE department_id = ?',
        (department_id,)
    ).fetchall()

    trainings = conn.execute(
        'SELECT * FROM training_programs WHERE department_id = ?',
        (department_id,)
    ).fetchall()

    attendance_records = conn.execute('''
        SELECT ar.id as record_id, ar.session_date, tp.training_name, 
               GROUP_CONCAT(e.name, ', ') as participants
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        LEFT JOIN attendance_details ad ON ar.id = ad.attendance_id
        LEFT JOIN employees e ON ad.employee_id = e.id
        WHERE tp.department_id = ?
        GROUP BY ar.id
        ORDER BY ar.session_date DESC
    ''', (department_id,)).fetchall()

    conn.close()

    return render_template('Department/attendance.html',
                           department=department,
                           employees=employees,
                           trainings=trainings,
                           attendance_records=attendance_records,
                           current_date=today)


@app.route('/export_attendance_period')
def export_attendance_period():
    department = request.args.get('department')
    month = request.args.get('month')
    year = request.args.get('year')


    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        flash('Department not found', 'error')
        return redirect(url_for('attendance'))

    department_id = dept['id']


    query_params = [department_id]
    date_filter = ""

    if month:
        date_filter += " AND strftime('%m', ar.session_date) = ?"
        query_params.append(f"{int(month):02d}")
    if year:
        date_filter += " AND strftime('%Y', ar.session_date) = ?"
        query_params.append(year)


    records = conn.execute(f'''
        SELECT ar.id, ar.session_date, tp.training_name, 
               GROUP_CONCAT(e.name, ', ') as participants,
               GROUP_CONCAT(e.designation, ', ') as designations,
               GROUP_CONCAT(e.employee_id, ', ') as employee_ids
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        LEFT JOIN attendance_details ad ON ar.id = ad.attendance_id
        LEFT JOIN employees e ON ad.employee_id = e.id
        WHERE tp.department_id = ? {date_filter}
        GROUP BY ar.id
        ORDER BY ar.session_date DESC
    ''', query_params).fetchall()
    conn.close()


    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Attendance Records')


    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})
    regular_format = workbook.add_format({'border': 1})


    headers = ['Training Name', 'Session Date', 'Participants', 'Employee IDs', 'Designations']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)


    for row, record in enumerate(records, start=1):
        worksheet.write(row, 0, record['training_name'], regular_format)
        worksheet.write(row, 1, record['session_date'], date_format)
        worksheet.write(row, 2, record['participants'], regular_format)
        worksheet.write(row, 3, record['employee_ids'], regular_format)
        worksheet.write(row, 4, record['designations'], regular_format)

    workbook.close()
    output.seek(0)


    filename = f"attendance_records_{department}"
    if year:
        filename += f"_{year}"
    if month:
        filename += f"_{month}"
    filename += ".xlsx"

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/export_all_attendance')
def export_all_attendance():
    department = request.args.get('department')
    month = request.args.get('month')
    year = request.args.get('year')


    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        flash('Department not found', 'error')
        return redirect(url_for('attendance'))

    department_id = dept['id']


    query_params = [department_id]
    date_filter = ""

    if month:
        date_filter += " AND strftime('%m', ar.session_date) = ?"
        query_params.append(f"{int(month):02d}")
    if year:
        date_filter += " AND strftime('%Y', ar.session_date) = ?"
        query_params.append(year)


    records = conn.execute(f'''
        SELECT ar.id, ar.session_date, tp.training_name, 
               GROUP_CONCAT(e.name, ', ') as participants,
               GROUP_CONCAT(e.designation, ', ') as designations,
               GROUP_CONCAT(e.employee_id, ', ') as employee_ids
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        LEFT JOIN attendance_details ad ON ar.id = ad.attendance_id
        LEFT JOIN employees e ON ad.employee_id = e.id
        WHERE tp.department_id = ? {date_filter}
        GROUP BY ar.id
        ORDER BY ar.session_date DESC
    ''', query_params).fetchall()
    conn.close()


    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Attendance Records')


    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})
    regular_format = workbook.add_format({'border': 1})


    headers = ['Training Name', 'Session Date', 'Participants', 'Employee IDs', 'Designations']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)


    for row, record in enumerate(records, start=1):
        worksheet.write(row, 0, record['training_name'], regular_format)
        worksheet.write(row, 1, record['session_date'], date_format)
        worksheet.write(row, 2, record['participants'], regular_format)
        worksheet.write(row, 3, record['employee_ids'], regular_format)
        worksheet.write(row, 4, record['designations'], regular_format)

    workbook.close()
    output.seek(0)


    filename = f"all_attendance_records_{department}.xlsx"

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@app.route('/export_attendance_record/<record_id>')
def export_attendance_record(record_id):
    if 'department_id' not in session:
        flash('Not authorized')
        return redirect(url_for('logout'))

    conn = get_db_connection()


    record = conn.execute('''
        SELECT ar.*, tp.training_name, d.name as department
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        JOIN departments d ON tp.department_id = d.id
        WHERE ar.id = ? AND tp.department_id = ?
    ''', (record_id, session['department_id'])).fetchone()

    if not record:
        conn.close()
        flash('Record not found')
        return redirect(url_for('attendance'))


    participants = conn.execute('''
        SELECT e.id, e.name, e.designation, e.employee_id
        FROM attendance_details ad
        JOIN employees e ON ad.employee_id = e.id
        WHERE ad.attendance_id = ?
    ''', (record_id,)).fetchall()
    conn.close()


    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Attendance Record')


    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm', 'border': 1})
    regular_format = workbook.add_format({'border': 1})
    bold_format = workbook.add_format({'bold': True, 'border': 1})


    worksheet.write(0, 0, 'Training Name:', bold_format)
    worksheet.write(0, 1, record['training_name'], regular_format)
    worksheet.write(1, 0, 'Session Date:', bold_format)
    worksheet.write(1, 1, record['session_date'], date_format)
    worksheet.write(2, 0, 'Department:', bold_format)
    worksheet.write(2, 1, record['department'], regular_format)


    worksheet.write(4, 0, 'Participants', header_format)
    worksheet.write(4, 1, 'Employee ID', header_format)
    worksheet.write(4, 2, 'Designation', header_format)


    for row, participant in enumerate(participants, start=5):
        worksheet.write(row, 0, participant['name'], regular_format)
        worksheet.write(row, 1, participant['employee_id'], regular_format)
        worksheet.write(row, 2, participant['designation'] or 'N/A', regular_format)

    workbook.close()
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=attendance_record_{record_id}.xlsx'
    return response


@app.route('/department_report')
def department_report():
    if session.get('role') not in ['ld_admin', 'dept_admin']:
        return redirect(url_for('logout'))

    department = request.args.get('department', session.get('department'))
    if not department:
        return redirect(url_for('logout'))

    conn = get_db_connection()


    stats = {
        'total_trainings': conn.execute(
            'SELECT COUNT(*) FROM training_programs WHERE department_id = ?',
            (session.get('department_id'),)
        ).fetchone()[0],
        'total_attendance_records': conn.execute('''
            SELECT COUNT(*) FROM attendance_records ar
            JOIN training_programs tp ON ar.training_id = tp.id
            WHERE tp.department_id = ?
        ''', (session.get('department_id'),)).fetchone()[0],
        'unique_employees_trained': conn.execute('''
            SELECT COUNT(DISTINCT ad.employee_id) 
            FROM attendance_details ad
            JOIN attendance_records ar ON ad.attendance_id = ar.id
            JOIN training_programs tp ON ar.training_id = tp.id
            WHERE tp.department_id = ?
        ''', (session.get('department_id'),)).fetchone()[0],
        'total_employees': conn.execute(
            'SELECT COUNT(*) FROM employees WHERE department_id = ?',
            (session.get('department_id'),)
        ).fetchone()[0]
    }

    trainings = []
    for training in conn.execute(
        'SELECT * FROM training_programs WHERE department_id = ? ORDER BY created_at DESC LIMIT 5',
        (session.get('department_id'),)
    ).fetchall():
        training_dict = dict(training)
        # Get assigned employees
        assigned_employees = conn.execute(
            'SELECT e.id, e.name FROM training_assignments ta '
            'JOIN employees e ON ta.employee_id = e.id '
            'WHERE ta.training_id = ?', (training['id'],)
        ).fetchall()
        training_dict['assigned_employees'] = [dict(emp) for emp in assigned_employees]
        trainings.append(training_dict)

    attendance_records = []
    for record in conn.execute('''
        SELECT ar.id, ar.session_date, tp.training_name
        FROM attendance_records ar
        JOIN training_programs tp ON ar.training_id = tp.id
        WHERE tp.department_id = ?
        ORDER BY ar.session_date DESC
        LIMIT 5
    ''', (session.get('department_id'),)).fetchall():
        record_dict = dict(record)
        # Get participants
        participants = conn.execute(
            'SELECT e.id, e.name FROM attendance_details ad '
            'JOIN employees e ON ad.employee_id = e.id '
            'WHERE ad.attendance_id = ?', (record['id'],)
        ).fetchall()
        record_dict['participants'] = [dict(emp) for emp in participants]
        attendance_records.append(record_dict)

    conn.close()

    return render_template('Department/report.html',
                         department=department,
                         trainings=trainings,
                         attendance_records=attendance_records,
                         stats=stats)


@app.route('/export_report')
@login_required()
def export_report():
    department = request.args.get('department', session.get('department'))
    month = request.args.get('month')
    year = request.args.get('year')

    if not department:
        return redirect(url_for('logout'))

    # Get department ID
    conn = get_db_connection()
    dept = conn.execute(
        'SELECT id FROM departments WHERE name = ?',
        (department,)
    ).fetchone()

    if not dept:
        conn.close()
        return redirect(url_for('logout'))

    department_id = dept['id']


    query_params = [department_id]
    date_filter = ""

    if month:
        date_filter += " AND strftime('%m', tp.created_at) = ?"
        query_params.append(f"{int(month):02d}")
    if year:
        date_filter += " AND strftime('%Y', tp.created_at) = ?"
        query_params.append(year)


    trainings = []
    for training in conn.execute(
            f'SELECT tp.* FROM training_programs tp '
            f'WHERE tp.department_id = ? {date_filter} '
            f'ORDER BY tp.created_at DESC',
            query_params
    ).fetchall():
        training_dict = dict(training)
        assigned_employees = conn.execute(
            'SELECT e.id, e.name, e.designation FROM training_assignments ta '
            'JOIN employees e ON ta.employee_id = e.id '
            'WHERE ta.training_id = ?', (training['id'],)
        ).fetchall()
        training_dict['assigned_employees'] = [dict(emp) for emp in assigned_employees]
        trainings.append(training_dict)

    attendance_records = []
    for record in conn.execute(
            f'SELECT ar.*, tp.training_name FROM attendance_records ar '
            f'JOIN training_programs tp ON ar.training_id = tp.id '
            f'WHERE tp.department_id = ? {date_filter} '
            f'ORDER BY ar.session_date DESC',
            query_params
    ).fetchall():
        record_dict = dict(record)
        # Get participants
        participants = conn.execute(
            'SELECT e.id, e.name, e.designation FROM attendance_details ad '
            'JOIN employees e ON ad.employee_id = e.id '
            'WHERE ad.attendance_id = ?', (record['id'],)
        ).fetchall()
        record_dict['participants'] = [dict(emp) for emp in participants]
        attendance_records.append(record_dict)

    conn.close()


    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)


    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#4472C4',
        'font_color': 'white',
        'border': 1
    })


    data_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })


    date_format = workbook.add_format({
        'num_format': 'yyyy-mm-dd hh:mm',
        'border': 1,
        'align': 'left'
    })


    training_sheet = workbook.add_worksheet('Training Programs')


    training_sheet.set_column('A:A', 30)
    training_sheet.set_column('B:B', 20)
    training_sheet.set_column('C:C', 15)
    training_sheet.set_column('D:D', 15)
    training_sheet.set_column('E:E', 15)
    training_sheet.set_column('F:F', 15)
    training_sheet.set_column('G:G', 40)


    training_sheet.write_row(0, 0, [
        'Training Name', 'Date Created', 'Assigned Employees',
        'Sessions Held', 'Unique Participants', 'Attendance Rate',
        'Training Details'
    ], header_format)


    for row, training in enumerate(trainings, start=1):
        # Calculate attendance stats
        assigned_count = len(training['assigned_employees'])


        conn = get_db_connection()
        session_count = conn.execute(
            'SELECT COUNT(*) FROM attendance_records WHERE training_id = ?',
            (training['id'],)
        ).fetchone()[0]


        participant_count = conn.execute(
            'SELECT COUNT(DISTINCT ad.employee_id) FROM attendance_details ad '
            'JOIN attendance_records ar ON ad.attendance_id = ar.id '
            'WHERE ar.training_id = ?', (training['id'],)
        ).fetchone()[0]
        conn.close()

        attendance_rate = 'N/A'
        if assigned_count > 0 and session_count > 0:
            rate = (participant_count / (assigned_count * session_count)) * 100
            attendance_rate = f"{round(rate)}%"

        training_data = json.loads(training['training_data'])
        details = '\n'.join([f"{k}: {v}" for k, v in training_data.items()
                             if k not in ['date', 'training_name', 'assigned_employees']])

        training_sheet.write(row, 0, training['training_name'], data_format)
        training_sheet.write(row, 1, training['created_at'], date_format)
        training_sheet.write_number(row, 2, assigned_count, data_format)
        training_sheet.write_number(row, 3, session_count, data_format)
        training_sheet.write_number(row, 4, participant_count, data_format)
        training_sheet.write(row, 5, attendance_rate, data_format)
        training_sheet.write(row, 6, details, data_format)


    employees_sheet = workbook.add_worksheet('Assigned Employees')


    employees_sheet.set_column('A:A', 30)  # Training Name
    employees_sheet.set_column('B:B', 25)  # Employee Name
    employees_sheet.set_column('C:C', 20)  # Designation
    employees_sheet.set_column('D:D', 15)  # Employee ID


    employees_sheet.write_row(0, 0, [
        'Training Name', 'Employee Name', 'Designation', 'Employee ID'
    ], header_format)


    row = 1
    for training in trainings:
        for emp in training['assigned_employees']:
            employees_sheet.write(row, 0, training['training_name'], data_format)
            employees_sheet.write(row, 1, emp['name'], data_format)
            employees_sheet.write(row, 2, emp['designation'], data_format)
            employees_sheet.write(row, 3, emp['id'], data_format)
            row += 1


    attendance_sheet = workbook.add_worksheet('Attendance Records')


    attendance_sheet.set_column('A:A', 30)  # Training Name
    attendance_sheet.set_column('B:B', 20)  # Session Date
    attendance_sheet.set_column('C:C', 15)  # Participants
    attendance_sheet.set_column('D:D', 15)  # Attendance Rate
    attendance_sheet.set_column('E:E', 40)  # Participant Names


    attendance_sheet.write_row(0, 0, [
        'Training Name', 'Session Date', 'Participants Count',
        'Attendance Rate', 'Participant Names'
    ], header_format)


    for row, record in enumerate(attendance_records, start=1):

        training = next((t for t in trainings if t['id'] == record['training_id']), None)
        assigned_count = len(training['assigned_employees']) if training else 0

        attendance_rate = 'N/A'
        if assigned_count > 0:
            rate = (len(record['participants']) / assigned_count) * 100
            attendance_rate = f"{round(rate)}%"

        participant_names = ', '.join([p['name'] for p in record['participants']])

        attendance_sheet.write(row, 0, record['training_name'], data_format)
        attendance_sheet.write(row, 1, record['session_date'], date_format)
        attendance_sheet.write_number(row, 2, len(record['participants']), data_format)
        attendance_sheet.write(row, 3, attendance_rate, data_format)
        attendance_sheet.write(row, 4, participant_names, data_format)

    summary_sheet = workbook.add_worksheet('Summary')

    summary_sheet.set_column('A:A', 30)
    summary_sheet.set_column('B:B', 20)

    summary_sheet.merge_range('A1:B1', f'{department} Department Training Report Summary', header_format)

    summary_data = [
        ['Total Training Programs', len(trainings)],
        ['Total Attendance Sessions', len(attendance_records)],
        ['Unique Employees Trained', len(set(
            emp['id'] for t in trainings for emp in t['assigned_employees']
        ))],
        ['Total Employees', len(set(
            emp['id'] for t in trainings for emp in t['assigned_employees']
        ))],
        ['Overall Attendance Rate', 'Calculated Rate']
    ]

    for row, (label, value) in enumerate(summary_data, start=2):
        summary_sheet.write(row, 0, label, header_format)
        summary_sheet.write(row, 1, value, data_format)

    workbook.close()
    output.seek(0)

    filename = f"{department}_training_report"
    if year:
        filename += f"_{year}"
    if month:
        filename += f"_{month}"
    filename += ".xlsx"

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


@app.route('/lnd_reports')
def lnd_reports():
    if session.get('role') != 'ld_admin':
        return redirect(url_for('logout'))


    conn = get_db_connection()
    departments = conn.execute('SELECT name FROM departments').fetchall()
    conn.close()

    department_names = [dept['name'] for dept in departments]

    return render_template('L&D/reports.html', departments=department_names)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('role_selection'))


if __name__ == '__main__':
    app.run(debug=True)